import requests, base64, json
import urllib.parse
from string import Template
from django.shortcuts import render
from django.http import HttpResponse
from .forms import UserForm
from datetime import datetime
import os
import logging
from openpyxl import Workbook


# Get an instance of a logger
logger = logging.getLogger(__name__)
# Build paths inside the project like this: os.path.join(BASE_DIR, ...)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


cookies = {}
headers={}

str_00 = 'bda88568a54f922fcdfc6dbf940e5d00'
str_0b = '56105c9ab348522591eea18fbe4d080b'
str_PNSESSIONID = 'PNSESSIONID'
unit = '1256 гап'
search_count = -9999

#####################################
def parse_file (name_file):
    dict_ = {}
    f = open(name_file, 'r')
    s = f.read()
    dict_={}
    list_ = s.splitlines()
    for item in list_:
        items = item.split(":")
        dict_[items[0]] = items[1].lstrip()
    return dict_
#####################################
def parse_json_file(json_file):
    with open(json_file) as json_file:
        return json.load(json_file)
#####################################
def make_str_cookie(cookies):
    str_cook = ''
    for key, value in cookies.items():
        str_cook += '{0}={1};'.format(key,value)
    return str_cook
#####################################
def checkType(value):
    if value is None:
        return ''
    else:
        return value

def getContent(military_unit, date_From, date_To):
    ##############################################
    #    Первый запрос - получаем 307 статус     #
    ##############################################
    html_string = '<html><table class="blueTable">'
    headers = parse_file(BASE_DIR+'/mu_files/mu_header1.txt')
    cookies = parse_file(BASE_DIR+'/mu_files/mu_cookie1.txt')
    url = 'https://pamyat-naroda.ru/'
    res1 = requests.get(url, allow_redirects=False)
    global search_count
    search_count = 0
    if(res1.status_code==307):
        #print('*********  1  **********')
        #print(res1.status_code)
        #print(res1.cookies[str_00])
        # Получили переменные из кук
        cookie_PNSESSIONID = res1.cookies['PNSESSIONID']
        cookie_00 = res1.cookies[str_00]
        cookie_0b = res1.cookies[str_0b]
        #####################################################
        # готовим 2-й запрос, посылаем с получанными куками #
        #####################################################
        #print('')
        #print('*********  2  **********')
        cookies = {}
        cookies[str_00]=cookie_00
        cookies[str_0b]=cookie_0b
        cookies[str_PNSESSIONID] = cookie_PNSESSIONID
        cookies['BITRIX_PN_DATALINE_LNG'] = 'ru'
        headers = parse_file(BASE_DIR+'/mu_files/mu_header2.txt')
        headers['Cookie'] = make_str_cookie(cookies)
        res2 = requests.get(url,cookies=cookies,headers=headers,allow_redirects = True)
        #######################################
        if(res2.status_code==200):
            #print(res2.status_code)
            #print(res2.cookies[str_00])
            ###########################################
            ##############   3-й запрос   #############
            ###########################################
            cookies = parse_json_file(BASE_DIR+'/mu_files/mu_cookie3.txt')
            cookies[str_00] = res2.cookies[str_00]
            cookies[str_0b] = res1.cookies[str_0b]
            cookies[str_PNSESSIONID] = res1.cookies[str_PNSESSIONID]
            cookies['r'] = res1.cookies[str_0b]
            headers = parse_file(BASE_DIR+'/mu_files/mu_header3.txt')
            headers['Cookie'] = make_str_cookie(cookies)
            headers['Content-Type'] = 'application/json'

            url3 = 'https://pamyat-naroda.ru/documents/'
            res3 = requests.get(url3,headers=headers,cookies=cookies)
            #print(res3.status_code)
            #print(res3.cookies[str_00])
            ############## 4-й запрос #############
            ############## 4-й запрос #############
            headers=parse_file(BASE_DIR+'/mu_files/mu_header4.txt')
            headers['Content-Type'] = 'application/json'
            headers['Origin']='https://pamyat-naroda.ru'
            headers['Referer']='https://pamyat-naroda.ru/documents/'

            bs = res3.cookies[str_00]
            bs += "=" * ((4 - len(res3.cookies[str_00]) % 4) % 4)
            bs = base64.b64decode(bs).decode()
            a_bs = bs.split('XXXXXX')[0]
            b_bs = bs.split('XXXXXX')[1].split('YYYYYY')[0]
            data_t = Template('{"query":{"bool":{"should":[{"bool":{"should":[{"match_phrase":{"document_type":"Боевые донесения, оперсводки"}},{"match_phrase":{"document_type":"Боевые приказы и распоряжения"}},{"match_phrase":{"document_type":"Отчеты о боевых действиях"}},{"match_phrase":{"document_type":"Переговоры"}},{"match_phrase":{"document_type":"Журналы боевых действий"}},{"match_phrase":{"document_type":"Директивы и указания"}},{"match_phrase":{"document_type":"Приказы"}},{"match_phrase":{"document_type":"Постановления"}},{"match_phrase":{"document_type":"Доклады"}},{"match_phrase":{"document_type":"Рапорты"}},{"match_phrase":{"document_type":"Разведывательные бюллетени и донесения"}},{"match_phrase":{"document_type":"Сведения"}},{"match_phrase":{"document_type":"Планы"}},{"match_phrase":{"document_type":"Планы операций"}},{"match_phrase":{"document_type":"Карты"}},{"match_phrase":{"document_type":"Схемы"}},{"match_phrase":{"document_type":"Справки"}},{"match_phrase":{"document_type":"Прочие документы"}}]}},{"bool":{"should":[{"bool":{"must":[{"range":{"date_from":{"lte":"${finish_date}"}}},{"range":{"date_to":{"gte":"${start_date}"}}}],"boost":3}},{"bool":{"must":[{"range":{"document_date_b":{"lte":"${finish_date}"}}},{"range":{"document_date_f":{"gte":"${start_date}"}}}],"boost":7}}]}},{"bool":{"should":[{"match_phrase":{"authors_list.keyword":{"query":"${military_unit}","boost":50}}},{"match":{"document_name":{"query":"${military_unit}","type":"phrase","boost":30}}},{"match":{"authors":{"query":"${military_unit}","type":"phrase","boost":20}}},{"match":{"army_unit_label.division":{"query":"${military_unit}","type":"phrase","boost":10}}},{"nested":{"path":"page_magazine","query":{"bool":{"must":[{"match":{"page_magazine.podrs":{"query":"${military_unit}","type":"phrase"}}},{"range":{"page_magazine.date_from":{"lte":"${finish_date}"}}},{"range":{"page_magazine.date_to":{"gte":"${start_date}"}}}]}},"boost":4}}]}}],"minimum_should_match":3}},"_source":["id","document_type","document_number","document_date_b","document_date_f","document_name","archive","fond","opis","delo","date_from","date_to","authors","geo_names","operation_name","secr","image_path","delo_id","deal_type","operation_name"],"size":"${size}","from":"${_from}"}')
            #data_t= Template('{"query":{"bool":{"should":[{"bool":{"should":[{"match_phrase":{"document_type":"Боевые донесения, оперсводки"}},{"match_phrase":{"document_type":"Боевые приказы и распоряжения"}},{"match_phrase":{"document_type":"Отчеты о боевых действиях"}},{"match_phrase":{"document_type":"Переговоры"}},{"match_phrase":{"document_type":"Журналы боевых действий"}},{"match_phrase":{"document_type":"Директивы и указания"}},{"match_phrase":{"document_type":"Приказы"}},{"match_phrase":{"document_type":"Постановления"}},{"match_phrase":{"document_type":"Доклады"}},{"match_phrase":{"document_type":"Рапорты"}},{"match_phrase":{"document_type":"Разведывательные бюллетени и донесения"}},{"match_phrase":{"document_type":"Сведения"}},{"match_phrase":{"document_type":"Планы"}},{"match_phrase":{"document_type":"Планы операций"}},{"match_phrase":{"document_type":"Карты"}},{"match_phrase":{"document_type":"Схемы"}},{"match_phrase":{"document_type":"Справки"}},{"match_phrase":{"document_type":"Прочие документы"}}]}},{"bool":{"should":[{"bool":{"must":[{"range":{"date_from":{"lte":"${finish_date}"}}},{"range":{"date_to":{"gte":"${start_date}"}}}],"boost":3}},{"bool":{"must":[{"range":{"document_date_b":{"lte":"${finish_date}"}}},{"range":{"document_date_f":{"gte":"${start_date}"}}}],"boost":7}}]}}],"minimum_should_match":2}},"_source":["id","document_type","document_number","document_date_b","document_date_f","document_name","archive","fond","opis","delo","date_from","date_to","authors","geo_names","operation_name","secr","image_path","delo_id","deal_type","operation_name"],"size":"${size}","from":"${para_from}"}')

            data_ = data_t.safe_substitute(start_date=date_From,finish_date=date_To, military_unit=military_unit,size=10,_from=0)
            url4 = 'https://cdn.pamyat-naroda.ru/data/'+a_bs+'/'+b_bs+'/pamyat/document,map,magazine/_search'
            res4 = requests.post(url4,data=data_.encode('utf-8'),headers=headers,allow_redirects = True)
            if(res4.status_code==200):
                data = json.loads(res4.text)
                total = data['hits']['total']
                search_count = total
                hits = data['hits']['hits']
                divisor = 100
                one, two = divmod(total,divisor)
                x=0
                count=0
                print('total = ',total)
                html_string += '<thead><tr><th>N</th><th>Тип документа</th><th>Содержание</th><th>Период</th><th>Авторы</th><th>Дата документа</th><th>Архив</th><th>Фонд</th><th>Опись</th><th>Дело</th><th>Док</th></tr></thead>'
                html_string += '<tbody>'
                table_string = Template('<tr><td>${cnt}</td><td>${col1}</td><td>${col2}</td><td>${col3}</td><td>${col4}</td><td>${col5}</td><td>${col6}</td><td>${col7}</td><td>${col8}</td><td>${col9}</td><td>${col10}</td></tr>')

                start=0
                stop=10
                step = 100
                for i in range(start,stop,step):
                    #print(i,start, stop, step)
                    data_ = data_t.safe_substitute(start_date=date_From,finish_date=date_To, military_unit=military_unit,size=step,_from=i)
                    #print(data)
                    url4 = 'https://cdn.pamyat-naroda.ru/data/'+a_bs+'/'+b_bs+'/pamyat/document,map,magazine/_search'
                    res4 = requests.post(url4,data=data_.encode('utf-8'),headers=headers,allow_redirects = True)
                    if(res4.status_code==200):
                        data = json.loads(res4.text)
                        hits = data['hits']['hits']
                        #search_count += len(hits)
                        result_array = []
                        for hit in hits:
                            src = hit['_source']
                            dict_record = {}
                            count+=1
                            data_string = table_string.safe_substitute(cnt=count,col1=checkType(src['document_type']),col2=checkType(src['document_name']),col3=checkType(src['date_from'])+'_'+checkType(src['date_to']),col4=checkType(src['authors']),col5=checkType(src['document_date_f']),col6=checkType(src['archive']),col7=checkType(src['fond']),col8=checkType(src['opis']),col9=checkType(src['delo']),col10='<a href=https://pamyat-naroda.ru/documents/view/?id='+hit['_id']+' target="_blank">Док</a>')
                            dict_record['cnt'] = count
                            dict_record['document_type'] = checkType(src['document_type'])
                            dict_record['document_name'] = checkType(src['document_name'])
                            dict_record['period'] = checkType(src['date_from'])+'_'+checkType(src['date_to'])
                            dict_record['authors'] = checkType(src['authors'])
                            dict_record['document_date_f'] = checkType(src['document_date_f'])
                            dict_record['archive'] = checkType(src['archive'])
                            dict_record['fond'] = checkType(src['fond'])
                            dict_record['opis'] = checkType(src['opis'])
                            dict_record['delo'] = checkType(src['delo'])
                            dict_record['link'] = 'https://pamyat-naroda.ru/documents/view/?id='+hit['_id']
                            result_array.append(dict_record)

                            html_string += data_string
                html_string+='</tbody></table></html>'
    return(html_string, result_array)

'''
                while(x< one*divisor):
                    data_ = data_t.safe_substitute(start_date=date_From,finish_date=date_To, military_unit=military_unit,size=divisor,para_from=x)
                    url4 = 'https://cdn.pamyat-naroda.ru/data/'+a_bs+'/'+b_bs+'/pamyat/document,map,magazine/_search'
                    res4 = requests.post(url4,data=data_.encode('utf-8'),headers=headers,allow_redirects = True)
                    if(res4.status_code==200):
                        data = json.loads(res4.text)
                        hits = data['hits']['hits']
                        search_count += len(hits)
                        for hit in hits:
                            src = hit['_source']
                            count+=1
                            logger.log(count)
                            data_string = table_string.safe_substitute(cnt=str(count),col1=src['document_type'],col2=src['document_name'],col3=src['date_from']+'-'+src['date_to'],col4=src['authors'],col5=src['document_date_f'],col6=src['archive'],col7=src['fond'],col8=src['opis'],col9=src['delo'],col10='<a href=https://pamyat-naroda.ru/documents/view/?id='+hit['_id']+' target="_blank">Док</a>')
                            html_string += data_string
                    x+=divisor

                data_ = data_t.safe_substitute(start_date=date_From,finish_date=date_To, military_unit=military_unit,size=two,para_from=x)
                url4 = 'https://cdn.pamyat-naroda.ru/data/'+a_bs+'/'+b_bs+'/pamyat/document,map,magazine/_search'
                res4 = requests.post(url4,data=data_.encode('utf-8'),headers=headers)
                if(res4.status_code==200):
                    data = json.loads(res4.text)
                    hits = data['hits']['hits']
                    search_count += len(hits)
                    for hit in hits:
                        src = hit['_source']
                        count+=1
                        data_string = table_string.safe_substitute(cnt=str(count),col1=src['document_type'],col2=src['document_name'],col3=src['date_from']+'-'+src['date_to'],col4=src['authors'],col5=src['document_date_f'],col6=src['archive'],col7=src['fond'],col8=src['opis'],col9=src['delo'],col10='<a href=https://pamyat-naroda.ru/documents/view/?id='+hit['_id']+' target="_blank">Док</a>')
                        html_string += data_string
            '''




def index(request):
    global search_count
    if request.method == "POST":

        unit = request.POST.get("unit")
        str_date_From = request.POST.get("date_From")
        str_date_To = request.POST.get("date_To")
        tmp_date_From = datetime.strptime(str_date_From,'%d.%m.%Y')
        tmp_date_To = datetime.strptime(str_date_To,'%d.%m.%Y')
        if(os.name == "nt"):
            date_From = tmp_date_From.strftime('%Y-%m-%d')
            date_To = tmp_date_To.strftime('%Y-%m-%d')
        else:
            date_From = tmp_date_From.strftime('%Y-%-m-%-d')
            date_To = tmp_date_To.strftime('%Y-%-m-%-d')
        # age = request.POST.get("age") # получение значения поля age
        #return HttpResponse("<h2>Hello, {0}</h2>".format(name))
        #return HttpResponse("<h2>date_From, {0}</h2>".format(date_From))
        #return HttpResponse(getContent(unit,date_From, date_To))
        userform = UserForm({'unit':unit,'date_From': str_date_From, 'date_To':str_date_To})
        table, arr = getContent(unit,date_From, date_To)
        if 'SaveData' in request.POST:
            #print(arr)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
                date=datetime.now().strftime('%Y-%m-%d'),
            )
            workbook = Workbook()
            
            # Get active worksheet/tab
            worksheet = workbook.active
            worksheet.title = 'Movies'
            columns = [
                'ID',
                'Тип документа',
                'Содержание',
                'Период',
                'Авторы',
                'Дата документа',
                'Архив',
                'Фонд',
                'Опист',
                'Дело',
                'Документ',
            ]
            row_num = 1

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title

            for rec in arr:
                row_num += 1
                row = [
                    rec.get('cnt'),
                    rec.get('document_type'),
                    rec.get('document_name'),
                    rec.get('period'),
                    rec.get('authors'),
                    rec.get('document_date_f'),
                    rec.get('archive'),
                    rec.get('fond'),
                    rec.get('opis'),
                    rec.get('delo'),
                    rec.get('link'),
                ]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
            workbook.save(response)
            return response
            #return
        else:
            return render(request, "search/index.html", {"form": userform,'search_count':'Найдено: '+str(search_count),"table_content": table})
    else:
        userform = UserForm()
        return render(request, "search/index.html", {"form": userform})
